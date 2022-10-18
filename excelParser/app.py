import openpyxl as xl


def process_workbook(original_filename, final_filename):
    workbook = xl.load_workbook(original_filename)

    # sheets = workbook.sheetnames
    sheet = workbook['Sheet1']

    corrected_cell_header = sheet.cell(1, 4)
    corrected_cell_header.value = "Correct Price"

    for row in range(2, sheet.max_row + 1):

        cell = sheet.cell(row, 3)
        corrected_value = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_value

    workbook.save(final_filename)


process_workbook("transactions.xlsx", "t2.xlsx")
